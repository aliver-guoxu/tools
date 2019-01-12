from xlrd import open_workbook
import openpyxl
import datetime
import time

class Excel_opration():
    def __init__(self):
        self.excel_path=input('请输入namelist文件目录：')
        self.template_excel=input('请输入模板文件的目录：')
        self.read_data={}
        self.all_data=[]
        self.final_read_data={}

    def read_excel(self):
        rb=open_workbook(self.excel_path)
        rs=rb.sheet_by_name('COP')
        num_cols=rs.ncols
        num_rows=rs.nrows
        for row in range(1,num_rows):
            for col in range(num_cols):
                self.read_data[rs.row_values(0)[col]]=rs.row_values(row)[col]
            self.final_read_data['Pers.No.']=self.read_data['Pers.No.']
            self.final_read_data['Name'] = self.read_data['Name']
            self.final_read_data['Annual salary'] = self.read_data['Annual salary']
            self.final_read_data['Target Bonus-Annual'] = self.read_data['Target Bonus-Annual']
            self.final_read_data['Ratio'] = self.read_data['Ratio']
            self.all_data.append(self.final_read_data)
            self.read_data={}
            self.final_read_data={}
        print(self.all_data)

    def prepare_data(self):
        for data in self.all_data:
            print(data)
            name=data['Name']
            pers_no=data['Pers.No.']
            file_name='Conversion-'+pers_no+'-'+name+'.XLSX'
            Annual_salary=data['Annual salary']
            Target_Bonus_Annual=data['Target Bonus-Annual']
            Ratio=data['Ratio']
            emplooy_id='Employee ID:  {}'.format(pers_no)
            employee_name='Employee Name:  {}'.format(name)
            time=self.date_time()
            data_time='Effective Date:  {}'.format(time)
            self.write_excel(file_name,emplooy_id,employee_name,data_time,Annual_salary,Target_Bonus_Annual,Ratio)

    def write_excel(self,filename,id,name,time,annual,target,ratio):
        wb=openpyxl.load_workbook(filename=self.template_excel,data_only=False)
        ws=wb.active
        ws.cell(row=3,column=1).value=id
        ws.cell(row=4,column=1).value=name
        ws.cell(row=5, column=1).value =time
        ws.cell(row=10, column=2).value = annual
        ws.cell(row=14, column=2).value = target
        ws.cell(row=14, column=3).value = ratio
        wb.save(filename)

    def date_time(self):
        data=datetime.datetime.now()
        data=str(data)[0:11]
        return data

    def run(self):
        self.read_excel()
        self.prepare_data()
        time.sleep(3)
        print('创建成功')

if __name__ == '__main__':
    excel_opration=Excel_opration()
    excel_opration.run()